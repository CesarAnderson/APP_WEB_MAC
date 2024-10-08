from flask import Flask, render_template, request, send_from_directory, url_for, jsonify
import os
import google.generativeai as genai
import fitz
from PIL import Image
from io import BytesIO
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import load_workbook
import subprocess
import platform


app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXCEL_FILE = 'Facturas.xlsx'

genai.configure(api_key="AIzaSyANGD5y7dErSN4FhDZdM0_rfUtBFB-8iR8")
model = genai.GenerativeModel("gemini-1.5-flash-001")

def convert_pdf_to_jpg(pdf_path):
    doc = fitz.open(pdf_path)
    page = doc[0]
    pix = page.get_pixmap(dpi=300)
    img_bytes = pix.tobytes("jpeg")
    return BytesIO(img_bytes)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'imagen' not in request.files:
            return jsonify({"error": "No se seleccionó ninguna imagen"}), 400

        imagen = request.files['imagen']
        if imagen.filename == '':
            return jsonify({"error": "No se seleccionó ninguna imagen"}), 400

        if imagen:
            filename = secure_filename(imagen.filename)
            file_extension = os.path.splitext(filename)[1].lower()
            
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            imagen.save(filepath)

            image_bytes = None
            if file_extension == '.pdf':
                image_bytes_io = convert_pdf_to_jpg(filepath)
                image_bytes = image_bytes_io.getvalue()
                jpg_filename = filename[:-4] + ".jpg"
                jpg_filepath = os.path.join(app.config['UPLOAD_FOLDER'], jpg_filename)
                with open(jpg_filepath, 'wb') as f:
                    f.write(image_bytes)
                filename = jpg_filename
            else:
                with open(filepath, 'rb') as f:
                    image_bytes = f.read()

            try:
                prompt = [
                    genai.protos.Part(text="Extrae el número de factura, la fecha de la factura en formato dd/mm/yyyy, el importe total, la base imponible y el valor del porcentaje del IVA (no quiero el importe del IVA, quiero el valor porcentual del iva). Devuelve los resultados separados(los valores numericos deben ser decimales separados por coma) por punto y coma en este orden: numerodefactura;fechadefactura;importetotal;baseimponible;iva"),
                    genai.protos.Part(inline_data=genai.protos.Blob(mime_type='image/jpeg', data=image_bytes))
                ]
                response = model.generate_content(prompt)
                descripcion = response.text.strip().replace("*", "")

                try:
                    numero_factura, fecha_factura, importe_total, base_imponible, iva = descripcion.split(";")
                    numero_factura = numero_factura.strip()[-5:].lstrip('0').replace('/', '').replace('-', '')
                    fecha_factura = fecha_factura.strip()
                    importe_total = importe_total.strip()
                    base_imponible = base_imponible.strip()
                    iva = iva.strip()
                except ValueError:
                    numero_factura = fecha_factura = importe_total = base_imponible = iva = "No encontrado"

                return jsonify({
                    "filename": filename,
                    "fecha_factura": fecha_factura,
                    "numero_factura": numero_factura,
                    "codigo_empresa": "",
                    "tipo_gasto": "",
                    "base_imponible": base_imponible,
                    "iva": iva,
                    "importe_total": importe_total,
                    "imageUrl": url_for('uploaded_file', filename=filename)
                })

            except Exception as e:
                return jsonify({
                    "error": f"Error al procesar la imagen: {str(e)}",
                    "filename": filename,
                    "imageUrl": url_for('uploaded_file', filename=filename)
                }), 500

    return render_template('index.html')

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/export-to-excel', methods=['POST'])
def export_to_excel():
    data = request.json
    
    # Load the existing Excel file or create a new one if it doesn't exist
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb['Hoja1']
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Hoja1'
        ws.append(['Fecha', 'Número', 'Codigo Empresa', 'Tipo de Gasto', 'Base Imponible', 'IVA', 'Importe', 'Archivo'])
    
    # Add new rows
    for row in data:
        ws.append([
            row['fecha_factura'],
            row['numero_factura'],
            row['codigo_empresa'],
            row['tipo_gasto'],
            row['base_imponible'],
            row['iva'],
            row['importe_total'],
            row['filename']
        ])
    
    # Save the workbook
    wb.save(EXCEL_FILE)
    
    return jsonify({"message": "Data exported successfully"}), 200

@app.route('/open-excel', methods=['GET'])
def open_excel():
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"success": False, "message": "Excel file does not exist"}), 404

    try:
        if platform.system() == "Windows":
            os.startfile(EXCEL_FILE)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(('open', EXCEL_FILE))
        else:  # Linux and other Unix-like
            subprocess.call(('xdg-open', EXCEL_FILE))
        return jsonify({"success": True, "message": "Excel file opened successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": f"Error opening Excel file: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)
